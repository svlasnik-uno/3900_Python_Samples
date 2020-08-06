import os
import openpyxl
from openpyxl import load_workbook

wb=load_workbook(filename = 'customers.xlsx',read_only=True)
ws = wb['customers']

for rowOfCellObjects in ws['A1':'G7']:
           for cellObj in rowOfCellObjects:
               print(cellObj.coordinate, cellObj.value)
           print('--- END OF ROW ---')
custlist =[]

for rowOfCellObjects in ws['A1':'G7']:
           for cellObj in rowOfCellObjects:
               custlist.append(cellObj.value)

print (custlist)

#for row in ws.rows:

#    for cell in row:
#        print(cell.value)
